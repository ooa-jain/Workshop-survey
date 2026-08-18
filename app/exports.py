"""
Spreadsheet exports.

The shared analysis page shows the per-student table on screen but leaves it
out of the printed PDF -- a class handout doesn't need a row per person. The
same table is offered as a real .xlsx instead, built here.

openpyxl is a pure-Python wheel (no system libraries), so this adds nothing to
the VPS beyond a pip install.
"""

import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

INK = "FF17131F"
LIME = "FFD6F53F"
WARN = "FFC3312B"
TEAL = "FF0E8F96"

HEAD_FILL = PatternFill("solid", fgColor=INK)
HEAD_FONT = Font(bold=True, color="FFFFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
LABEL_FONT = Font(bold=True)


def safe_filename(text, fallback="analysis", ext="xlsx"):
    """A filename safe to drop into a Content-Disposition header: no quotes,
    no newlines, no path separators, and never empty."""
    cleaned = re.sub(r"[^A-Za-z0-9 _-]+", " ", str(text or ""))
    cleaned = re.sub(r"\s+", " ", cleaned).strip()[:60].strip()
    return f"{cleaned or fallback}.{ext}"


def _write_header(ws, row, headers):
    for col, head in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=head)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    return row + 1


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def shared_analysis_xlsx(share_title, group_label, o, people, window_label=""):
    """
    share_title / group_label: strings shown on the shared page.
    o:       the build_outcome() dict.
    people:  the same anonymisation-applied rows the page renders, so the
             spreadsheet can never reveal a name the page itself hides, and
             carries no email addresses at all.
    """
    wb = Workbook()

    # --- Sheet 1: every student ---------------------------------------------
    ws = wb.active
    ws.title = "Students"

    ws["A1"] = share_title
    ws["A1"].font = TITLE_FONT
    scope = f"{group_label} — Pre → Post Survey 1"
    if window_label:
        scope += f" — counting only students who filled {window_label}"
    ws["A2"] = scope
    ws["A2"].font = Font(italic=True, size=10)
    ws["A3"] = "Job Intelligence is scored 0–100. Shift is Post Survey 1 minus Pre."
    ws["A3"].font = Font(size=10, color="FF6B6377")

    row = _write_header(ws, 5, [
        "#", "Student", "Role at Pre", "Role at Post Survey 1", "Changed role",
        "Job Intelligence (Pre)", "Job Intelligence (Post Survey 1)", "Shift",
    ])
    for i, p in enumerate(people, start=1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=p["who"])
        ws.cell(row=row, column=3, value=p["quad_pre"])
        ws.cell(row=row, column=4, value=p["quad_now"])
        ws.cell(row=row, column=5, value="Yes" if p["moved"] else "No")
        ws.cell(row=row, column=6, value=p["ji_pre"])
        ws.cell(row=row, column=7, value=p["ji_now"])
        shift = ws.cell(row=row, column=8, value=p["delta"])
        shift.font = Font(bold=True, color=(TEAL if p["delta"] > 0 else
                                            (WARN if p["delta"] < 0 else "FF6B6377")))
        shift.number_format = "+0.0;-0.0;0"
        row += 1

    _autosize(ws, [5, 26, 20, 22, 14, 22, 30, 10])
    ws.freeze_panes = "A6"

    # --- Sheet 2: the group summary -----------------------------------------
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = share_title
    ws2["A1"].font = TITLE_FONT
    ws2["A2"] = group_label
    ws2["A2"].font = Font(italic=True, size=10)

    row = 4
    for label, value in [
        ("Students in this analysis", o["n"]),
        ("Job Intelligence — Pre (mean)", o["ji_pre_mean"]),
        ("Job Intelligence — Post Survey 1 (mean)", o["ji_now_mean"]),
        ("Mean shift", o["ji_delta"]),
        ("Rose", o["rose"]),
        ("Fell", o["fell"]),
        ("No change", o["same"]),
        ("Changed role (quadrant)", o["n_moved_quad"]),
    ]:
        ws2.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws2.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    ws2.cell(row=row, column=1, value="Quadrant population").font = TITLE_FONT
    row += 1
    row = _write_header(ws2, row, ["Quadrant", "Pre", "Post Survey 1", "Change"])
    for q in o["quadrants"]:
        before, after = o["quad_counts_pre"][q], o["quad_counts_now"][q]
        ws2.cell(row=row, column=1, value=q)
        ws2.cell(row=row, column=2, value=before)
        ws2.cell(row=row, column=3, value=after)
        change = ws2.cell(row=row, column=4, value=after - before)
        change.number_format = "+0;-0;0"
        row += 1

    _autosize(ws2, [40, 16, 18, 12])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
